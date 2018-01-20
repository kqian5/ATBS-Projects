#! python3
# updateProduce.py - Updates costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb.get_sheet_by_name("Sheet")

PRICES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICES:
        sheet.cell(row=rowNum, column=2).value = PRICES[produceName]

wb.save("updatedProduceSales.xlsx")


